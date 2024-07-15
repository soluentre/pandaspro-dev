from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string
import re


class CellPro:
    def __init__(self, cell: str):
        if re.fullmatch(r'[A-Z]+\d+', cell) or re.fullmatch(r'[A-Z]+\d+:[A-Z]+\d+', cell):
            self.valid = True
            if ':' in cell:
                self.celltype = 'range'
                self.cell_start = cell.split(':')[0].strip()
                self.cell_stop = cell.split(':')[1].strip()
                self.width = cell_index(self.cell_stop)[1] - cell_index(self.cell_start)[1] + 1
                self.height = cell_index(self.cell_stop)[0] - cell_index(self.cell_start)[0] + 1
            else:
                self.celltype = 'cell'
                self.cell_cal = cell
            self.cell = cell
        else:
            self.valid = False

    @property
    def cell_index(self):
        if self.celltype == 'cell':
            return cell_index(self.cell_cal)
        else:
            raise ValueError('range object does not have index_cell property')

    def resize(self, row_resize, col_resize):
        if self.celltype == 'cell':
            return CellPro(resize(self.cell_cal, row_resize, col_resize))
        else:
            return CellPro(resize(self.cell_start, row_resize, col_resize))

    def resize_w(self, col_resize):
        if self.celltype == 'cell':
            return self.resize(1, col_resize)
        else:
            bottom_left = offset(self.cell_stop, 0, -(self.width - 1))
            bottom_right = offset(bottom_left, 0, col_resize - 1)
            return CellPro(self.cell_start + ':' + bottom_right)

    def resize_h(self, row_resize):
        if self.celltype == 'cell':
            return self.resize(row_resize, 1)
        else:
            top_right = offset(self.cell_stop, -(self.height - 1), 0)
            bottom_right = offset(top_right, row_resize - 1, 0)
            return CellPro(self.cell_start + ':' + bottom_right)

    def offset(self, down_offset, right_offset):
        if self.celltype == 'cell':
            return CellPro(offset(self.cell, down_offset, right_offset))
        else:
            newstart = offset(self.cell_start, down_offset, right_offset)
            newstop = offset(self.cell_stop, down_offset, right_offset)
            newrange = newstart + ':' + newstop
            return CellPro(newrange)

    def range(self, down_offset, right_offset, row_resize, col_resize):
        return self.offset(down_offset, right_offset).resize(row_resize, col_resize)


def index_cell(row_index, column_index):
    return get_column_letter(column_index) + str(row_index)


def cell_index(cell: str) -> list:
    """
    This function converts an Excel cell name (e.g., 'A1') into its corresponding row and column indices using
    openpyxl's utility functions. It separates the alphabetic column identifier(s) and the numeric row identifier,
    then converts the column identifier to a numeric index.

    Parameters
    ----------
    cell : str
        The name of the cell in Excel format (e.g., 'A1', 'B22'), where letters refer to the column and numbers
        refer to the row.

    Returns
    -------
    tuple
        A tuple containing two elements:
        - The first element is an integer representing the row number.
        - The second element is an integer representing the column number.

    Notes
    -----
    The function relies on openpyxl's `coordinate_from_string` to split the cell name into its letter and number
    components, and `column_index_from_string` to convert the column letter(s) to a numeric index.
    It assumes the input is a valid Excel cell reference.

    Examples
    --------
    >>> cell_index('C3')
    This would return (3, 3), indicating that the cell is in the 3rd row and 3rd column of the spreadsheet.
    """

    column_letter, row_number = coordinate_from_string(cell)  # Separate letter and number
    column_number = column_index_from_string(column_letter)  # Convert letter to number
    return [row_number, column_number]


def resize(
        cell: str,
        row_resize: int,
        col_resize: int
) -> str:
    """
    Adjusts the size of a cell range starting from a specified cell in Excel format, by adding a specified number
    of rows and columns to it. The function returns the new cell range in Excel notation. It uses the `index_cell`
    function to get the row and column indices of the starting cell, then calculates the ending cell's indices
    based on the resize parameters.

    Parameters
    ----------
    cell : str
        The starting cell's name in Excel format (e.g., 'A1'), where letters represent the column and numbers
        represent the row.
    row_resize : int
        The number of rows to add to the starting cell's row to determine the ending cell's row. The resize value
        includes the starting row itself.
    col_resize : int
        The number of columns to add to the starting cell's column to determine the ending cell's column.
        The resize value includes the starting column itself.

    Returns
    -------
    str
        The new cell range in Excel format (e.g., 'A1:B2'), where the first part ('A1') is the starting cell,
        and the second part ('B2') is the ending cell determined by the resize parameters.

    Notes
    -----
    The function assumes the provided cell name and resize parameters are valid. The ending cell is calculated
    by adding `row_resize - 1` to the starting row and `col_resize - 1` to the starting column, accounting for
    the inclusion of the starting cell in the resize count.

    Examples
    --------
    >>> resize('B2', 3, 2)
    This would return 'B2:C4', indicating that starting from cell 'B2', the new range extends 3 rows down and 2
    columns to the right, ending at cell 'C4'.
    """
    row, col = cell_index(cell)
    new_row = row + row_resize - 1
    new_col = col + col_resize - 1
    start_column_letter = get_column_letter(col)
    end_column_letter = get_column_letter(new_col)
    result = f"{start_column_letter}{row}:{end_column_letter}{new_row}"
    if new_row <= 0 or new_col <= 0:
        raise ValueError(f"Excel min row is 0 and min col is A, the result would be invalid {result}")
    return result


def offset(
        cell: str,
        down_offset: int,
        right_offset: int
) -> str:
    """
    Calculates the Excel cell reference offset from a given starting cell by a specified number of rows and columns.
    This function allows for moving a cell reference vertically and horizontally based on the provided offsets.

    Parameters
    ----------
    cell : str
        The starting cell's name in Excel format (e.g., 'A1'), where letters represent the column and numbers
        represent the row.
    down_offset : int
        The number of rows to move down from the starting cell. A positive value moves the cell reference down,
        while a negative value moves it up.
    right_offset : int
        The number of columns to move right from the starting cell. A positive value moves the cell reference to
        the right, while a negative value moves it to the left.

    Returns
    -------
    str
        The new cell reference in Excel format (e.g., 'B2') after applying the given row and column offsets.

    Raises
    ------
    ValueError
        If the resulting cell reference is outside the valid Excel sheet range, specifically if the row or
        column index is less than 1.

    Notes
    -----
    The function checks for invalid resulting indices, ensuring that the new cell reference does not exceed Excel's
    minimum row and column limits. Excel sheets start at row 1 and column 'A'.

    Examples
    --------
    >>> offset('A1', 2, 3)
    This would return 'D3', indicating that starting from cell 'A1', moving 2 rows down and 3 columns to the
    right lands at cell 'D3'.
    """
    row, col = cell_index(cell)
    new_row = row + down_offset
    new_col = col + right_offset
    new_column_letter = get_column_letter(new_col)
    if new_row <= 0 or new_col <= 0:
        raise ValueError(
            f"Excel min row is 0 and min col is A, the result would be invalid {new_column_letter}{new_row}")
    return f"{new_column_letter}{new_row}"


def getrange(
        cell: str,
        down_offset,
        right_offset,
        row_resize,
        col_resize
 ) -> str:
    return resize(offset(cell, down_offset, right_offset), row_resize, col_resize)


def get_cell_lists(rowlist: list,
                   columnlist: list,
                   orientation: str = 'c') -> dict:
    """
    Generates a dictionary of cell lists from specified row and column ranges, organized by either rows or columns
    based on the given orientation. Each key in the dictionary corresponds to a list of cell references in Excel format.

    Parameters
    ----------
    rowlist : list
        A list of row indices (integers) to be included in the cell lists.
    columnlist : list
        A list of column letters (strings) corresponding to the columns to be included.
    orientation : str, optional
        Specifies the orientation for grouping cell references. 'r' for row-wise grouping, 'c' for
        column-wise grouping (default is 'c').

    Returns
    -------
    dict
        A dictionary where each key ('cells0', 'cells1', ...) maps to a list of cell references in
        the specified orientation.

    Examples
    --------
    >>> get_cell_lists([1, 2], ['A', 'B'], 'r')
    This will return a dictionary with row-wise grouped cell references, e.g.
    {'cells0': ['A1', 'B1'], 'cells1': ['A2', 'B2']}.
    """
    result_dict = {}
    i = 0

    if orientation == 'r':
        for row in rowlist:
            temp_list = []
            for col in columnlist:
                temp_list.append(col + str(row))
            result_dict[f'cells{i}'] = temp_list
            i += 1

    elif orientation == 'c':
        for col in columnlist:
            temp_list = []
            for row in rowlist:
                temp_list.append(col + str(row))
            result_dict[f'cells{i}'] = temp_list
            i += 1

    return result_dict


def cell_combine_by_row(cells: list = None):
    cells_by_row = {}
    for cell in cells:
        row = cell_index(cell)[0]
        column = get_column_letter(cell_index(cell)[1])
        if row not in cells_by_row:
            cells_by_row[row] = []
        cells_by_row[row].append(column)

    merged_cells = {}
    for row, columns in cells_by_row.items():
        merged = []
        current_range = [columns[0]]

        for i in range(1, len(columns)):
            if cell_index(columns[i]+'1')[1] == cell_index(current_range[-1]+'1')[1] + 1:
                current_range.append(columns[i])
            else:
                merged.append(f"{current_range[0]}{row}:{current_range[-1]}{row}")
                current_range = [columns[i]]

        merged.append(f"{current_range[0]}{row}:{current_range[-1]}{row}")
        merged_cells[row] = merged

    return merged_cells


def cell_combine_by_column(cells: list = None):
    cells_by_column = {}
    for cell in cells:
        row = cell_index(cell)[0]
        column = get_column_letter(cell_index(cell)[1])
        if column not in cells_by_column:
            cells_by_column[column] = []
        cells_by_column[column].append(row)

    merged_cells = {}
    for column, rows in cells_by_column.items():
        merged = []
        current_range = [rows[0]]

        for i in range(1, len(rows)):
            if rows[i] == current_range[-1] + 1:
                current_range.append(rows[i])
            else:
                merged.append(f"{column}{current_range[0]}:{column}{current_range[-1]}")
                current_range = [rows[i]]

        merged.append(f"{column}{current_range[0]}:{column}{current_range[-1]}")
        merged_cells[column] = merged

    return merged_cells


if __name__ == '__main__':
    print(cell_combine_by_row('E16,F16,E21,E22,E25,E26,E27,E50,E56'.split(',')))