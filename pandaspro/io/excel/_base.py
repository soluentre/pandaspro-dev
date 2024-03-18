import re
import pandas as pd
from openpyxl.utils import column_index_from_string, get_column_letter
from pandaspro.core.frame import cFrame
from pandaspro.io.excel._utils import lowervarlist


def pwread(
        file: str,
        sheet_name: str | int = 0,
        cellrange: str = None,
        firstrow: bool = True,
        skiprows: int = None,
        returnmap: bool = True,
        debug: bool = False,
        **kwargs
):
    # Decide the file type and call the right function
    if file.endswith('.xlsx') or file.endswith('.xlsm'):
        func = pd.read_excel
        filetype = 'excel'
    elif file.endswith('.csv'):
        func = lambda f, **kwargs: pd.read_csv(f, **{k: v for k, v in kwargs.items() if k != 'sheet_name'})
        filetype = 'csv'
    else:
        raise TypeError('Only support .xlsx/.xlsm and .csv')

    # Define get_columns_between (will return a list of columns)
    def get_columns_between(start, end):
        start_index = column_index_from_string(start)
        end_index = column_index_from_string(end)
        columns = []
        for i in range(start_index, end_index + 1):
            columns.append(get_column_letter(i))
        return columns

    # Main function: reading accordingly as the input from the cell parameter
    if cellrange:
        if re.match(r"^[A-Z]+\d+:[A-Z]+\d+$", cellrange):
            match = re.match(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$", cellrange)
            start_col, start_row, end_col, end_row = match.group(1), int(match.group(2)), match.group(3), int(match.group(4))
            usecols_arg = f"{start_col}:{end_col}" if filetype == 'excel' else [i for i in range(column_index_from_string(start_col) - 1, column_index_from_string(end_col))]
            if firstrow:
                dfresult = func(file, sheet_name=sheet_name, usecols=usecols_arg, header=start_row - 1, nrows=end_row - start_row, **kwargs)
            if not firstrow:
                df = func(file, sheet_name=sheet_name, usecols=usecols_arg, header=None, skiprows=start_row - 1, nrows=end_row - start_row + 1, **kwargs)
                df.columns = get_columns_between(start_col, end_col)
                dfresult = df

        elif re.match(r"^[A-Z]+:[A-Z]+$", cellrange):
            match = re.match(r"^([A-Z]+):([A-Z]+)$", cellrange)
            start_col, end_col = match.group(1), match.group(2)
            usecols_arg = f"{start_col}:{end_col}" if filetype == 'excel' else [i for i in range(column_index_from_string(start_col) - 1, column_index_from_string(end_col))]
            if firstrow:
                dfresult = func(file, sheet_name=sheet_name, usecols=usecols_arg, **kwargs)
            if not firstrow:
                df = func(file, sheet_name=sheet_name, usecols=usecols_arg, header=None, **kwargs)
                df.columns = get_columns_between(start_col, end_col)
                dfresult = df

        elif re.match(r"^\d+:\d+$", cellrange):
            match = re.match(r"^(\d+):(\d+)$", cellrange)
            start_row, end_row = int(match.group(1)), int(match.group(2))
            if firstrow:
                dfresult = func(file, sheet_name=sheet_name, header=start_row - 1, nrows=end_row - start_row, **kwargs)
            if not firstrow:
                df = func(file, sheet_name=sheet_name, header=None, skiprows=start_row - 1, nrows=end_row - start_row + 1, **kwargs)
                df.columns = get_columns_between("A", get_column_letter(df.shape[1]))
                dfresult = df

        elif re.match(r"^[A-Z]+(,[A-Z]+)*$", cellrange):
            usecols_arg = cellrange if filetype == 'excel' else [column_index_from_string(letter.strip()) - 1 for letter in cellrange.split(',')]
            # Since in this scenario, the inputs are A, B, D, etc.
            # And therefore skiprows parameter is used to declare the starting point for each/all columns
            if firstrow:
                df = func(file, sheet_name=sheet_name, usecols=usecols_arg, skiprows=skiprows, **kwargs)
                dfresult = df
            if not firstrow:
                df = func(file, sheet_name=sheet_name, usecols=usecols_arg, header=None, skiprows=skiprows, **kwargs)
                df.columns = [get_column_letter(num + 1) for num in df.columns]
                dfresult = df

        elif re.match(r"^[A-Z]+\d+$", cellrange):
            match = re.match(r"^([A-Z]+)(\d+)$", cellrange)
            start_col, start_row = match.group(1), int(match.group(2))
            end_row, end_col = func(file, sheet_name=sheet_name, header=None, **kwargs).shape[0], \
                get_column_letter(func(file, sheet_name=sheet_name, header=None, **kwargs).shape[1])
            usecols_arg = f"{start_col}:{end_col}" if filetype == 'excel' else [i for i in range(
                column_index_from_string(start_col) - 1, column_index_from_string(end_col))]
            if firstrow:
                df = func(file, sheet_name=sheet_name, usecols=usecols_arg, header=start_row - 1, nrows=end_row - start_row, **kwargs)
                dfresult = df
            if not firstrow:
                df = func(file, sheet_name=sheet_name, usecols=usecols_arg, header=None, skiprows=start_row - 1,
                          nrows=end_row - start_row + 1, **kwargs)
                df.columns = get_columns_between(start_col, end_col)
                dfresult = df

        else:
            raise ValueError('Format of cell para. is not valid')

    else:
        if firstrow:
            dfresult = func(file, sheet_name=sheet_name, **kwargs)
        if not firstrow:
            dfresult = func(file, sheet_name=sheet_name, header=None, **kwargs)

    if returnmap:
        return cFrame(lowervarlist(dfresult)), lowervarlist(dfresult, 'revert')
    else:
        return cFrame(lowervarlist(dfresult))
