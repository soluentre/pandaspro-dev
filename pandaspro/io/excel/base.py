import re
from typing import Union

import numpy as np
import pandas as pd
from openpyxl.utils import column_index_from_string, get_column_letter
from pandaspro.core.frame import FramePro


def pwread(
        file: str,
        sheet_name: str | int = 0,
        cellrange: str = None,
        firstrow: bool = True,
        skiprows: int = None,
        returnmap: bool = True,
        debug: bool = False,
        **kwargs
):
    # Decide the file type and call the right function
    if file.endswith('.xlsx') or file.endswith('.xlsm'):
        func = pd.read_excel
        filetype = 'excel'
    elif file.endswith('.csv'):
        func = lambda f, **kwargs: pd.read_csv(f, **{k: v for k, v in kwargs.items() if k != 'sheet_name'})
        filetype = 'csv'
    else:
        raise TypeError('Only support .xlsx/.xlsm and .csv')

    # Define get_columns_between (will return a list of columns)
    def get_columns_between(start, end):
        start_index = column_index_from_string(start)
        end_index = column_index_from_string(end)
        columns = []
        for i in range(start_index, end_index + 1):
            columns.append(get_column_letter(i))
        return columns

    # Main function: reading accordingly as the input from the cell parameter
    if cellrange:
        if re.match(r"^[A-Z]+\d+:[A-Z]+\d+$", cellrange):
            match = re.match(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$", cellrange)
            start_col, start_row, end_col, end_row = match.group(1), int(match.group(2)), match.group(3), int(match.group(4))
            usecols_arg = f"{start_col}:{end_col}" if filetype == 'excel' else [i for i in range(column_index_from_string(start_col) - 1, column_index_from_string(end_col))]
            if firstrow:
                dfresult = func(file, sheet_name=sheet_name, usecols=usecols_arg, header=start_row - 1, nrows=end_row - start_row, **kwargs)
            if not firstrow:
                df = func(file, sheet_name=sheet_name, usecols=usecols_arg, header=None, skiprows=start_row - 1, nrows=end_row - start_row + 1, **kwargs)
                df.columns = get_columns_between(start_col, end_col)
                dfresult = df

        elif re.match(r"^[A-Z]+:[A-Z]+$", cellrange):
            match = re.match(r"^([A-Z]+):([A-Z]+)$", cellrange)
            start_col, end_col = match.group(1), match.group(2)
            usecols_arg = f"{start_col}:{end_col}" if filetype == 'excel' else [i for i in range(column_index_from_string(start_col) - 1, column_index_from_string(end_col))]
            if firstrow:
                dfresult = func(file, sheet_name=sheet_name, usecols=usecols_arg, **kwargs)
            if not firstrow:
                df = func(file, sheet_name=sheet_name, usecols=usecols_arg, header=None, **kwargs)
                df.columns = get_columns_between(start_col, end_col)
                dfresult = df

        elif re.match(r"^\d+:\d+$", cellrange):
            match = re.match(r"^(\d+):(\d+)$", cellrange)
            start_row, end_row = int(match.group(1)), int(match.group(2))
            if firstrow:
                dfresult = func(file, sheet_name=sheet_name, header=start_row - 1, nrows=end_row - start_row, **kwargs)
            if not firstrow:
                df = func(file, sheet_name=sheet_name, header=None, skiprows=start_row - 1, nrows=end_row - start_row + 1, **kwargs)
                df.columns = get_columns_between("A", get_column_letter(df.shape[1]))
                dfresult = df

        elif re.match(r"^[A-Z]+(,[A-Z]+)*$", cellrange):
            usecols_arg = cellrange if filetype == 'excel' else [column_index_from_string(letter.strip()) - 1 for letter in cellrange.split(',')]
            # Since in this scenario, the inputs are A, B, D, etc.
            # And therefore skiprows parameter is used to declare the starting point for each/all columns
            if firstrow:
                df = func(file, sheet_name=sheet_name, usecols=usecols_arg, skiprows=skiprows, **kwargs)
                dfresult = df
            if not firstrow:
                df = func(file, sheet_name=sheet_name, usecols=usecols_arg, header=None, skiprows=skiprows, **kwargs)
                df.columns = [get_column_letter(num + 1) for num in df.columns]
                dfresult = df

        elif re.match(r"^[A-Z]+\d+$", cellrange):
            match = re.match(r"^([A-Z]+)(\d+)$", cellrange)
            start_col, start_row = match.group(1), int(match.group(2))
            end_row, end_col = func(file, sheet_name=sheet_name, header=None, **kwargs).shape[0], \
                get_column_letter(func(file, sheet_name=sheet_name, header=None, **kwargs).shape[1])
            usecols_arg = f"{start_col}:{end_col}" if filetype == 'excel' else [i for i in range(
                column_index_from_string(start_col) - 1, column_index_from_string(end_col))]
            if firstrow:
                df = func(file, sheet_name=sheet_name, usecols=usecols_arg, header=start_row - 1, nrows=end_row - start_row, **kwargs)
                dfresult = df
            if not firstrow:
                df = func(file, sheet_name=sheet_name, usecols=usecols_arg, header=None, skiprows=start_row - 1,
                          nrows=end_row - start_row + 1, **kwargs)
                df.columns = get_columns_between(start_col, end_col)
                dfresult = df

        else:
            raise ValueError('Format of cell para. is not valid')

    else:
        if firstrow:
            dfresult = func(file, sheet_name=sheet_name, **kwargs)
        if not firstrow:
            dfresult = func(file, sheet_name=sheet_name, header=None, **kwargs)

    if returnmap:
        return FramePro(lowervarlist(dfresult)), lowervarlist(dfresult, 'revert')
    else:
        return FramePro(lowervarlist(dfresult))


def lowervarlist(
        data,
        engine: str = 'data',
        inplace: bool = False
) -> Union[pd.DataFrame, list, dict]:
    """
    This function renames the columns of a DataFrame by formatting the original column names
    according to a specified pattern, primarily to ensure the column names are
    valid identifiers suitable for use in queries or further processing.
    The function provides options for the type of output through the `engine` parameter
    and can perform the operation in-place if desired.

    Parameters
    ----------
    data : DataFrame
        The DataFrame whose columns are to be renamed.
    engine : str, optional
        Determines the type of output returned by the function. Options include:
        - 'data': Returns a new DataFrame with renamed columns (default).
        - 'column': Returns a list of new column names.
        - 'update_map': Returns a dictionary mapping original column names to new column names.
        - 'revert_map': Returns a dictionary mapping new column names back to original column names.
    inplace : bool, optional
        If True, the column renaming is applied in-place, and the function returns None. Defaults to False.

    Returns
    -------
    DataFrame, list, dict, or None
        The return type depends on the `engine` parameter:
        - If `engine` is 'data', returns a new DataFrame with renamed columns.
        - If `engine` is 'column', returns a list of the new column names.
        - If `engine` is 'update_map', returns a dictionary mapping from original to new column names.
        - If `engine` is 'revert_map', returns a dictionary mapping from new to original column names.
        - If `inplace` is True, the function modifies the input DataFrame in-place and returns None.

    Notes
    -----
    The function formats column names by replacing non-alphanumeric characters with underscores, converting to lowercase, and appending a suffix to duplicate names to ensure uniqueness.

    Examples
    --------
    >>> mydf = pd.DataFrame(np.random.rand(3, 3), columns=['Column 1', 'Column-2', 'Column 3'])
    >>> lowervarlist(mydf, 'update_map')
    This will return a dictionary mapping the original column names to their new, formatted names, such as {'Column 1': 'column_1', 'Column-2': 'column_2', 'Column 3': 'column_3'}.
    """

    _engines = {
        'data': 0,
        'columns': 1,
        'update': 2,
        'revert': 3
    }

    # Get the original list of column names
    oldname = data.columns.to_list()
    pattern = re.compile('\W+')

    # Dictionary to track the occurrence of each formatted column name
    name_count = {}
    newname = []
    for name in oldname:
        # Format the column name
        formatted_name = re.sub(pattern, '_', str(name)).lower().strip("_")

        # Increment count and modify name if it's a duplicate
        if formatted_name in name_count:
            name_count[formatted_name] += 1
            formatted_name += f"_{name_count[formatted_name]}"
        else:
            name_count[formatted_name] = 0

        newname.append(formatted_name)

    # Create a mapping of old column names to new column names
    mapping_update = {old: new for old, new in zip(oldname, newname)}
    mapping_revert = {new: old for new, old in zip(newname, oldname)}

    # Rename the columns in the DataFrame
    df = data.rename(columns=mapping_update)
    cols = df.columns.to_list()

    if inplace:
        data.rename(columns=mapping_update, inplace=True)
    else:
        return [df, cols, mapping_update, mapping_revert][_engines[engine]]
